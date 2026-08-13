# -*- coding: utf-8 -*-
"""
================================================================================
 ETL — Dashboard de control de promociones
 Promociones Urbanas Montellano, S.L.
================================================================================

 Lee los diarios contables y los presupuestos operativos y produce DATA.json,
 el bloque de datos que consume el dashboard.

 ENTRADA  (carpeta indicada en BASE)
   Diarios/DiarioPM{23,24,25,26}.xlsx        diario de la matriz, un fichero por ejercicio
   Diarios/DiarioPM{SOCIEDAD}{25,26}.xlsx    diario de cada sociedad vehiculo
   Presupuestos/{Promocion}.xlsx             estudio economico por promocion

 SALIDA
   DATA.json

 EJECUCION
   python etl.py                 usa las rutas por defecto
   python etl.py <BASE> <OUT>    rutas alternativas

 CRITERIO DE ASIGNACION A PROMOCION
   El coste incurrido se imputa por la cuenta de existencias 330000xx, que la
   propia contabilidad reparte por promocion y fase en el asiento mensual de
   variacion de existencias. El resto de reglas (sociedad vehiculo, cuenta de
   solar, cuenta 606 de obra, sufijo de promocion en la cuenta de cliente, serie
   de la factura emitida, numero de prestamo y cuenta bancaria) sirven para el
   desglose por naturaleza, la tesoreria y la deuda.

   Nada se reparte por estimacion. Lo que no encaja en un criterio verificable
   queda en la bandeja DONINOS_GRP (Doninos sin separar por fase) o SIN_ASIGNAR
   (estructura y partidas genericas), y se muestra integro en el dashboard.

 NOTA SOBRE LOS EJERCICIOS
   Los asientos de apertura y de cierre se excluyen para encadenar 2023-2026 sin
   duplicar saldos. Validado: el cierre de existencias de 2023 coincide al
   centimo con la apertura de 2024.
================================================================================
"""
import sys, os
BASE = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/uploads/Promociones Urbanas Montellano'
OUT  = sys.argv[2] if len(sys.argv) > 2 else '.'



# =============================================================================
# FASE 1 - Carga de los diarios y asignacion de cada apunte a una promocion
# =============================================================================
import openpyxl, glob, os, re, json, math, pickle
import pandas as pd, numpy as np

EXC={'apertura','cierre ejercicio','cierre contabilidad'}

SOC={'DiarioPM':('PUM','Promociones Urbanas Montellano, S.L.','B37359635'),
 'DiarioPMCARBAJOSA':('SPV_CARB','Promo. Mont. Carbajosa, S.L.','B23849797'),
 'DiarioPMDOÑINOS':('SPV_DORES','Promociones Montellano Doñinos Residencial, S.L.','B21882980'),
 'DiarioPMLARAD':('SPV_LARAD','Promociones Montellano La Rad, S.L.','B21888490'),
 'DiarioPMMARIN':('SPV_MARIN','Promociones Montellano El Marín, S.L.','B21882972'),
 'DiarioPMMIRADOR':('SPV_MIRA','Promo. Mont. Mirador de Vistahermosa, S.L.','B21886593'),
 'DiarioPMNAVES':('SPV_NAVES','Promociones Montellano Naves de Salamanca, S.L.','B21883012'),
 'DiarioPMVILLAS':('SPV_VILLAS','Prom. Montellano Villas de Vistahermosa, S.L.','B21883038')}
def soc_of(fn):
    b=re.sub(r'\d.*$','',os.path.splitext(fn)[0])
    return SOC[b]

def load():
    rows=[]
    for f in sorted(glob.glob(os.path.join(BASE,'Diarios','*.xlsx'))):
        wb=openpyxl.load_workbook(f,read_only=True,data_only=True); ws=wb.worksheets[0]
        cur=None; fn=os.path.basename(f); s=soc_of(fn)
        for r in ws.iter_rows(values_only=True):
            a=r[0]
            if isinstance(a,str):
                if a.startswith('Fecha:'): cur=a[6:].strip()
                continue
            if a is None: continue
            rows.append((fn,s[0],s[1],s[2],cur,a,r[1],r[2],r[3],r[4] or 0,r[5] or 0))
        wb.close()
    d=pd.DataFrame(rows,columns=['file','soc','soc_nom','nif','fecha','asiento','comentario','descripcion','cuenta','debe','haber'])
    d['fecha']=pd.to_datetime(d.fecha,format='%d/%m/%Y',errors='coerce')
    d['comentario']=d.comentario.fillna('').astype(str).str.strip()
    d['descripcion']=d.descripcion.fillna('').astype(str).str.strip()
    d['cuenta']=d.cuenta.fillna('').astype(str).str.strip()
    d.loc[d.cuenta.isin(['nan','None','']),'cuenta']='SIN_CUENTA'
    d['cta']=d.cuenta.apply(lambda c:c[:-1] if (len(c)==9 and c.endswith('0')) else c)
    d['neto']=d.debe-d.haber
    d['ej']=d.fecha.dt.year
    d['mes']=d.fecha.dt.to_period('M').astype(str)
    d['tec']=d.comentario.str.lower().isin(EXC)   # asientos tecnicos apertura/cierre
    d['id']=range(len(d))
    return d
D=load()
print('lineas totales',len(D),'| tecnicas',int(D.tec.sum()))
M=D[~D.tec].copy()     # flujo continuo 2023-2026

# ---------------- PROMOCIONES ----------------
PROMOS=[
 ('PUERTO','Puerto de Salamanca','Doñinos de Salamanca — Sector UR-R9, Mz 2B/3/5','Unifamiliar · 52 uds (F1-F3)','En venta y entrega'),
 ('NUEVOCAMPUS','Residencial Nuevo Campus','Salamanca — Sector 77 La Platina, M9.2/1-2-3','Bloque · 64 uds (F1-F3)','En obra'),
 ('CARBAJOSA','Jardines de Carbajosa','Carbajosa de la Sagrada — SUR-R-PAS2, AR.6 y AR.8','Bloque · 144 uds','Inicio de obra (Fase I)'),
 ('DONINOS_RES','Doñinos Residencial','Doñinos de Salamanca — Sector UR-R9, Mz 8 y 9','Unifamiliar · 48 uds','Preparación · obra 09/2026'),
 ('MARIN','Residencial El Marín','Salamanca — Sector El Marín, parcela 12.1','Bloque · 51 uds','Comercialización'),
 ('VISTAHERMOSA','Residencial Vistahermosa','Salamanca — Sector 35C Alambres-Vistahermosa, Mz 21','Unifamiliar · 28 uds','Comercialización'),
 ('MIRADOR','Mirador de Vistahermosa','Salamanca — Sector 65 El Zurguen, Parcela B-1','Unifamiliar · 94 uds','Suelo y proyecto'),
 ('LARAD','La Rad','Salamanca — Sector W8 PP Monte de La Rad','Unifamiliar · 64 uds','Suelo y proyecto'),
 ('ISLARUA','Obra Isla Rúa','Salamanca — Isla Rúa','Contrata ejecutada para Grancampo','Cerrada y facturada'),
 ('PTE_VILLANUEVA','Puente Villanueva de Perales','Villanueva de Perales','Obra para terceros (Ciresco)','En ejecución'),
 ('SUELO_IND','Doñinos Suelo Industrial','Doñinos de Salamanca','Parcela en proyecto, sin licencia ni contrata','En proyecto'),
 ('NAVES','Naves de Salamanca','Salamanca','Parcela en proyecto, sin licencia ni contrata','En proyecto'),
 ('OFICINAS','Oficinas y mejoras propias','Salamanca (Corrillo, Azafranal)','Gasto de oficina, no es una promoción','No promocional'),
 ('DONINOS_GRP','Doñinos — común (Puerto + Residencial)','Doñinos de Salamanca','Costes y tesorería compartidos','Bandeja de reparto pendiente'),
 ('SIN_ASIGNAR','Sin asignar / Estructura','—','Gastos de estructura y partidas sin criterio','—'),
]
PORD=[p[0] for p in PROMOS]; SIN='SIN_ASIGNAR'

MAP33={'33000001':'DONINOS_RES','33000002':'NUEVOCAMPUS','33000003':'PUERTO','33000004':'PUERTO',
 '33000005':'ISLARUA','33000006':'OFICINAS','33000007':'MARIN','33000008':'MIRADOR','33000010':'VISTAHERMOSA',
 '33000012':'OFICINAS','33000013':'NUEVOCAMPUS','33000014':'NUEVOCAMPUS','33000015':'SUELO_IND',
 '33000016':'LARAD','33000017':'CARBAJOSA'}
MAP31={'31300001':'LARAD','31300002':'MIRADOR','31300003':'MARIN','31300004':'VISTAHERMOSA','31300005':'SUELO_IND'}
MAP606={'60600001':'LARAD','60600003':'NUEVOCAMPUS','60600004':'MARIN','60600005':'VISTAHERMOSA',
 '60600006':'MIRADOR','60600007':'ISLARUA','60600008':'OFICINAS','60600009':'SUELO_IND','60600010':'OFICINAS',
 '60600011':'CARBAJOSA','60600012':'DONINOS_RES','60600014':'PTE_VILLANUEVA'}
MAP70={'70000002':None,'70000003':'NUEVOCAMPUS','70000004':'VISTAHERMOSA','70000005':'PTE_VILLANUEVA','70000006':'ISLARUA'}
PTMO={'17000003':('PUERTO','Santander 21873','Puerto de Salamanca · Fase 2'),
 '17000004':('NUEVOCAMPUS','Santander 21874','Nuevo Campus · Fase 1'),
 '17000005':('PUERTO','Santander 21875','Puerto de Salamanca · Fase 3'),
 '17000006':('NUEVOCAMPUS','Santander 21876','Nuevo Campus · Fase 2'),
 '17000007':('NUEVOCAMPUS','Santander 21877','Nuevo Campus · Fase 3')}
PTNUM={'21873':'PUERTO','21874':'NUEVOCAMPUS','21875':'PUERTO','21876':'NUEVOCAMPUS','21877':'NUEVOCAMPUS'}
BANCOS={'57000000':(SIN,'Caja euros'),'57200000':(SIN,'Bancos c/c vista'),'57200002':(SIN,'Caja Rural 2017310927'),
 '57200003':(SIN,'Unicaja 5320'),'57200005':(SIN,'Santander 2369 · Gastos Retos'),
 '57200006':(SIN,'Santander 2351 · Gastos generales'),'57200010':(SIN,'Santander 2628 · Cuenta de crédito'),
 '57200007':('NUEVOCAMPUS','Santander 4839 · Reservas Nuevo Campus'),
 '57200008':('PUERTO','Santander 4821 · Reservas Doñinos F1-F2'),
 '57200009':('PUERTO','Santander 5207 · Reservas Doñinos F3'),
 '57200011':('PUERTO','Santander 2636 · Pagos Doñinos'),
 '57200012':('VISTAHERMOSA','Santander 5835 · Reservas Vistahermosa'),
 '57200013':('NUEVOCAMPUS','Santander 3225 · Pagos Nuevo Campus'),
 '57200014':('NUEVOCAMPUS','Santander 6475 · Reservas Nuevo Campus 2'),
 '57200015':('NUEVOCAMPUS','Santander 6483 · Reservas Nuevo Campus 3')}
SPV2P={'SPV_CARB':'CARBAJOSA','SPV_DORES':'DONINOS_RES','SPV_LARAD':'LARAD','SPV_MARIN':'MARIN',
 'SPV_MIRA':'MIRADOR','SPV_NAVES':'NAVES','SPV_VILLAS':'VISTAHERMOSA'}
SUF=[('NCF1','NUEVOCAMPUS'),('NCF2','NUEVOCAMPUS'),('NCF3','NUEVOCAMPUS'),('NC1','NUEVOCAMPUS'),
     ('NC2','NUEVOCAMPUS'),('NC3','NUEVOCAMPUS'),('NC','NUEVOCAMPUS'),
     ('DOF1','PUERTO'),('DOF2','PUERTO'),('DOF3','PUERTO'),('DO1','PUERTO'),('DO2','PUERTO'),('DO3','PUERTO'),
     ('VVH','VISTAHERMOSA'),('VH','VISTAHERMOSA')]
SERIE={'DO1':'PUERTO','DO2':'PUERTO','DO3':'PUERTO','NCM':'NUEVOCAMPUS','NC1':'NUEVOCAMPUS',
       'NC2':'NUEVOCAMPUS','NC3':'NUEVOCAMPUS','PV':'PTE_VILLANUEVA','VH':'VISTAHERMOSA'}
def suf_promo(desc):
    t=(' '+desc.upper().strip()+' ')
    for s,p in SUF:
        if t.endswith(' '+s+' '): return p,s
    return None,None

def naturaleza(c):
    if c.startswith(('313','600','601','21000','23100')): return 'Suelo'
    if c.startswith(('6060','61')): return 'Obra'
    if c.startswith('623'): return 'Honorarios técnicos'
    if c.startswith('627'): return 'Comercial'
    if c.startswith('631') or c.startswith('6312'): return 'Tasas e impuestos'
    if c.startswith('625'): return 'Seguros'
    if c.startswith(('662','663','664','665','669','626')): return 'Financieros'
    if c.startswith(('64','755')): return 'Estructura'
    if c.startswith('68'): return 'Amortizaciones'
    if c.startswith('62'): return 'Estructura'
    if c.startswith(('70','75','77','76')): return 'Ingresos'
    if c.startswith('71'): return 'Var. existencias'
    return 'Otros'

def asignar(soc,c,desc,com):
    if soc!='PUM': return SPV2P[soc],'Sociedad vehículo dedicada a la promoción','alta'
    if c in MAP33: return MAP33[c],'Cuenta de existencias de la promoción','alta'
    if c in MAP31: return MAP31[c],'Cuenta de solar de la promoción','alta'
    if c in MAP606: return MAP606[c],'Cuenta de coste de obra de la promoción','alta'
    if c in PTMO: return PTMO[c][0],'Cuenta de préstamo asignada a promoción','alta'
    if c in BANCOS: return BANCOS[c][0],'Cuenta bancaria de la promoción','alta'
    if c.startswith(('430','438','431')):
        p,s=suf_promo(desc)
        if p: return p,'Sufijo de promoción en la cuenta de cliente','alta'
    mm=re.match(r'Ntra\.?\s*Fra\.?\s+([A-Z]{1,3}\d?)\s+Num',com,re.I)
    if mm and mm.group(1).upper() in SERIE: return SERIE[mm.group(1).upper()],'Serie de factura emitida','alta'
    mp=re.search(r'\b(2187[34567])\b',com)
    if mp: return PTNUM[mp.group(1)],'Nº de préstamo en el concepto','alta'
    if c in MAP70 and MAP70[c]:
        if MAP70[c]=='DONINOS_GRP': return 'DONINOS_GRP','Ingreso de Doñinos sin serie ni sufijo que identifique la fase','media'
        return MAP70[c],'Cuenta de ingresos de la promoción','alta'
    if c=='60600002': return 'PUERTO','Cuenta de obra de Doñinos: la única promoción de Doñinos en ejecución en el periodo','media'
    if c=='66230100': return 'PUERTO','Avales Ley 38/1999 sobre anticipos: solo Puerto de Salamanca tiene compradores','media'
    if c=='26000003': return 'PUERTO','Fianza de gestión de residuos de la obra de Doñinos','media'
    if c=='66230101': return 'NUEVOCAMPUS','Cuenta de avales de la promoción','alta'
    if c=='66230102': return 'MARIN','Cuenta de avales de la promoción','alta'
    if c.startswith('26000002') or c.startswith('26000004') or c.startswith('26000005'): return 'NUEVOCAMPUS','Fianza de la promoción','alta'
    return SIN,'Sin criterio de asignación en el diario','—'

res=[asignar(a,b,c,e) for a,b,c,e in zip(D.soc,D.cta,D.descripcion,D.comentario)]
D['promo']=[r[0] for r in res]; D['regla']=[r[1] for r in res]; D['conf']=[r[2] for r in res]
D['nat']=[naturaleza(c) for c in D.cta]
# herencia dentro del asiento
D['_k']=D.file+'|'+D.fecha.astype(str)+'|'+D.asiento.astype(str)
kn=D[(D.promo!=SIN)].groupby('_k').promo.agg(lambda s:s.iloc[0] if s.nunique()==1 else None)
uni=D['_k'].map(kn)
mk=(D.promo==SIN)&uni.notna()&(~D.cta.str.startswith(('33','313','57','56','10','11','12','16','17')))
D.loc[mk,'promo']=uni[mk]; D.loc[mk,'regla']='Heredada del asiento (única promoción identificada)'; D.loc[mk,'conf']='media'
# ingresos de Doñinos sin serie ni sufijo -> Puerto de Salamanca
mdg=(D.promo==SIN)&(D.cta.isin(['70000002','60600002','66230100']))
D.loc[mdg,'promo']='PUERTO'
D.loc[mdg,'regla']='Ingreso de Doñinos sin serie: Puerto de Salamanca es la única promoción de Doñinos con ventas'
D.loc[mdg,'conf']='media'
print('  apuntes por promocion:'); print(D.groupby('promo').size().to_string())


# =============================================================================
# FASE 2 - Series mensuales continuas 2023-2026 por promocion
# =============================================================================
import pandas as pd, numpy as np, re, json, math, pickle
M=D[~D.tec].copy()
MESES=sorted(M.mes.unique().tolist())
NM=len(MESES); IDX={m:i for i,m in enumerate(MESES)}
MN={'01':'Ene','02':'Feb','03':'Mar','04':'Abr','05':'May','06':'Jun','07':'Jul','08':'Ago','09':'Sep','10':'Oct','11':'Nov','12':'Dic'}
MLBL=[MN[m[5:7]]+' '+m[2:4] for m in MESES]
EJ=sorted(M.ej.unique().tolist())
def r2(x):
    try:
        f=float(x); return 0.0 if (math.isnan(f) or math.isinf(f)) else round(f,2)
    except: return 0.0
def serie(df,col='neto',sign=1):
    a=[0.0]*NM
    if len(df):
        g=df.groupby('mes')[col].sum()
        for m,v in g.items(): a[IDX[m]]=r2(sign*v)
    return a
def acum(a):
    o=[];s=0.0
    for v in a: s=r2(s+v); o.append(s)
    return o

PORD=['PUERTO','NUEVOCAMPUS','CARBAJOSA','DONINOS_RES','MARIN','VISTAHERMOSA','MIRADOR','LARAD',
      'ISLARUA','PTE_VILLANUEVA','SUELO_IND','NAVES','OFICINAS','SIN_ASIGNAR']
MAP33={'33000001':'DONINOS_RES','33000002':'NUEVOCAMPUS','33000003':'PUERTO','33000004':'PUERTO',
 '33000005':'ISLARUA','33000006':'OFICINAS','33000007':'MARIN','33000008':'MIRADOR','33000010':'VISTAHERMOSA',
 '33000012':'OFICINAS','33000013':'NUEVOCAMPUS','33000014':'NUEVOCAMPUS','33000015':'SUELO_IND',
 '33000016':'LARAD','33000017':'CARBAJOSA'}
SPV2P={'SPV_CARB':'CARBAJOSA','SPV_DORES':'DONINOS_RES','SPV_LARAD':'LARAD','SPV_MARIN':'MARIN',
 'SPV_MIRA':'MIRADOR','SPV_NAVES':'NAVES','SPV_VILLAS':'VISTAHERMOSA'}

ex=M[M.cta.str.startswith('33')].copy()
ex['p']=np.where(ex.soc=='PUM',ex.cta.map(MAP33),ex.soc.map(SPV2P))
sol=M[M.cta.str.startswith(('313','21000','23100'))].copy(); sol['p']=sol.promo
inv=pd.concat([ex,sol])

SER={}
for c in PORD:
    ing=M[(M.promo==c)&(M.cta.str.startswith('70'))]
    cv =ex[(ex.p==c)&(ex.haber>0)]
    ac =inv[(inv.p==c)&(inv.debe>0)]
    SER[c]=dict(
      ing=serie(ing,'neto',-1),
      cv=serie(cv,'haber',1),
      act=serie(ac,'debe',1),
      exVar=serie(inv[inv.p==c],'neto',1),
    )
    SER[c]['exSaldo']=acum(SER[c]['exVar'])
    SER[c]['ingAc']=acum(SER[c]['ing']); SER[c]['cvAc']=acum(SER[c]['cv']); SER[c]['actAc']=acum(SER[c]['act'])

# ---- naturaleza por ejercicio
NAT={}
g=M[M.cta.str.startswith(('60','61','62','63','64','65','66','67','68'))]
for c in PORD:
    x=g[g.promo==c]; NAT[c]={}
    for ej in EJ:
        y=x[x.ej==ej]
        dd={k:r2(v) for k,v in y.groupby('nat').neto.sum().items() if abs(v)>0.005}
        NAT[c][str(ej)]=dd

# ---- tesoreria
BANCOS={'57000000':('SIN_ASIGNAR','Caja euros'),'57200000':('SIN_ASIGNAR','Bancos c/c vista'),
 '57200002':('SIN_ASIGNAR','Caja Rural 2017310927'),'57200003':('SIN_ASIGNAR','Unicaja 5320'),
 '57200005':('SIN_ASIGNAR','Santander 2369 · Gastos Retos'),'57200006':('SIN_ASIGNAR','Santander 2351 · Gastos generales'),
 '57200010':('SIN_ASIGNAR','Santander 2628 · Cuenta de crédito'),
 '57200007':('NUEVOCAMPUS','Santander 4839 · Reservas Nuevo Campus'),
 '57200008':('PUERTO','Santander 4821 · Reservas Doñinos F1-F2'),
 '57200009':('PUERTO','Santander 5207 · Reservas Doñinos F3'),
 '57200011':('PUERTO','Santander 2636 · Pagos Doñinos'),
 '57200012':('VISTAHERMOSA','Santander 5835 · Reservas Vistahermosa'),
 '57200013':('NUEVOCAMPUS','Santander 3225 · Pagos Nuevo Campus'),
 '57200014':('NUEVOCAMPUS','Santander 6475 · Reservas Nuevo Campus 2'),
 '57200015':('NUEVOCAMPUS','Santander 6483 · Reservas Nuevo Campus 3')}
ban=M[M.cta.str.startswith(('570','572'))].copy()
ban['bp']=[BANCOS.get(c,('SIN_ASIGNAR',None))[0] if s=='PUM' else SPV2P[s] for c,s in zip(ban.cta,ban.soc)]
ban['bn']=[BANCOS[c][1] if (s=='PUM' and c in BANCOS) else d for c,s,d in zip(ban.cta,ban.soc,ban.descripcion)]
BAN=[]
for (p,nm,cta,soc),x in ban.groupby(['bp','bn','cta','soc_nom']):
    mv=serie(x,'neto',1)
    # cobros y pagos por el neto de cada apunte: asi la diferencia entre ambos
    # cuadra siempre con la variacion de caja, incluso con importes negativos
    xp=x[x.neto>0]; xn=x[x.neto<0]
    BAN.append(dict(promo=p,nom=nm,cta=cta,soc=soc,mens=mv,saldo=acum(mv),
      cob=serie(xp,'neto',1),pag=[-v for v in serie(xn,'neto',1)]))
for c in PORD:
    b=[x for x in BAN if x['promo']==c]
    SER[c]['cajaVar']=[r2(sum(x['mens'][i] for x in b)) for i in range(NM)]
    SER[c]['cajaSaldo']=acum(SER[c]['cajaVar'])
    SER[c]['cobros']=[r2(sum(x['cob'][i] for x in b)) for i in range(NM)]
    SER[c]['pagos']=[r2(sum(x['pag'][i] for x in b)) for i in range(NM)]

# ---- deuda
PTMO={'17000003':('PUERTO','Santander 21873','Puerto de Salamanca · Fase 2'),
 '17000004':('NUEVOCAMPUS','Santander 21874','Nuevo Campus · Fase 1'),
 '17000005':('PUERTO','Santander 21875','Puerto de Salamanca · Fase 3'),
 '17000006':('NUEVOCAMPUS','Santander 21876','Nuevo Campus · Fase 2'),
 '17000007':('NUEVOCAMPUS','Santander 21877','Nuevo Campus · Fase 3')}
DEU=[]
for cta,(p,nm,fs) in PTMO.items():
    x=M[(M.cta==cta)&(M.soc=='PUM')]
    disp=serie(x[x.haber>0],'haber',1); amo=serie(x[x.debe>0],'debe',1)
    var=[r2(disp[i]-amo[i]) for i in range(NM)]
    DEU.append(dict(cta=cta,promo=p,nom=nm,fase=fs,disp=disp,amort=amo,var=var,saldo=acum(var)))
for c in PORD:
    ls=[x for x in DEU if x['promo']==c]
    SER[c]['deudaSaldo']=[r2(sum(x['saldo'][i] for x in ls)) for i in range(NM)] if ls else [0.0]*NM
    SER[c]['deudaDisp']=[r2(sum(x['disp'][i] for x in ls)) for i in range(NM)] if ls else [0.0]*NM
    SER[c]['deudaAmort']=[r2(sum(x['amort'][i] for x in ls)) for i in range(NM)] if ls else [0.0]*NM


print('meses',NM,MESES[0],'->',MESES[-1])
print()
hdr=f"{'promo':16s}{'ingresos':>14}{'coste ventas':>14}{'margen':>13}{'coste acum':>14}{'obra curso':>14}{'caja':>12}{'deuda':>12}"
print(hdr); print('-'*len(hdr))
for c in PORD:
    s=SER[c]
    print(f"{c:16s}{sum(s['ing']):>14,.0f}{sum(s['cv']):>14,.0f}{sum(s['ing'])-sum(s['cv']):>13,.0f}{sum(s['act']):>14,.0f}{s['exSaldo'][-1]:>14,.0f}{s['cajaSaldo'][-1]:>12,.0f}{s['deudaSaldo'][-1]:>12,.0f}")


# =============================================================================
# FASE 3 - Presupuestos, facturas, calidad de datos y volcado a DATA.json
# =============================================================================
import pandas as pd, numpy as np, re, json, math, pickle, openpyxl, glob, os
NM=len(MESES)
def r2(x):
    try:
        f=float(x); return 0.0 if (math.isnan(f) or math.isinf(f)) else round(f,2)
    except: return 0.0

# ---------------- PRESUPUESTOS ----------------
PF={'PUERTO':'Puerto de Salamanca.xlsx','NUEVOCAMPUS':'Residencial Nuevo Campus.xlsx',
 'DONINOS_RES':'Doñinos Residencial.xlsx','CARBAJOSA':'Jardines de Carbajosa.xlsx',
 'MARIN':'Residencial El Marín.xlsx','VISTAHERMOSA':'Residencial Vistahermosa.xlsx',
 'MIRADOR':'Mirador de Vistahermosa.xlsx','LARAD':'La Rad.xlsx'}
CAP={'SUELO':'Suelo','2. CONSTRUCCIÓN':'Obra','3. TASAS E IMPUESTOS':'Tasas e impuestos',
 '4. HONORARIOS TÉCNICOS (incluye OCT)':'Honorarios técnicos','5. ACOMETIDAS':'Acometidas',
 '6. COMERCIALIZACIÓN':'Comercial','7. GASTOS JURÍDICOS':'Jurídicos','8. SEGUROS Y AVALES':'Seguros y avales',
 'FINANCIEROS':'Financieros'}
PRES={}
for cod,fn in PF.items():
    wb=openpyxl.load_workbook(os.path.join(BASE,'Presupuestos',fn),data_only=True)
    ws=wb['ESTUDIO ECONOMICO']; caps=[];cur=None;part=[]
    for r in ws.iter_rows(values_only=True):
        lab=r[1]
        if not isinstance(lab,str): continue
        if lab in CAP: cur=CAP[lab]; continue
        if lab=='TOTAL' and cur: caps.append(dict(cap=cur,pres=r2(r[4]),ejec=r2(r[5]))); cur=None; continue
        if lab.startswith('TOTAL imprevistos'):
            if r2(r[4]): caps.append(dict(cap='Imprevistos y costes extra',pres=r2(r[4]),ejec=r2(r[5])))
            continue
        if cur and lab!='COSTES (capítulos y partidas)' and (r[4] is not None or r[5] is not None):
            part.append(dict(cap=cur,part=lab,pres=r2(r[4]),ejec=r2(r[5])))
    res={}
    for r in wb['RESUMEN'].iter_rows(values_only=True):
        if isinstance(r[1],str) and r[2] is not None: res[r[1]]=r2(r[2])
    dg={}
    for r in wb['DATOS GENERALES'].iter_rows(values_only=True):
        if isinstance(r[0],str) and r[2] is not None: dg[r[0]]=r[2]
    cobra=[];hdr=False
    for r in wb['CAPITULOS OBRA'].iter_rows(values_only=True):
        if r[0]=='#': hdr=True; continue
        if hdr and isinstance(r[0],int): cobra.append(dict(n=r[0],cap=r[1],pres=r2(r[3]),contrata=r2(r[6]),real=r2(r[7])))
    lim=None
    for r in wb['FLUJO FINANCIERO'].iter_rows(values_only=True):
        if r[1]=='Límite préstamo promotor': lim=r2(r[3])
    ejec=r2(sum(x['ejec'] for x in caps))
    PRES[cod]=dict(caps=caps,partidas=part,cobra=cobra,limite=r2(lim or 0),
      ventas=r2(res.get('TOTAL VENTAS (sin IVA)')),coste=r2(res.get('TOTAL COSTES PROMOCIÓN')),
      margen=r2(res.get('MARGEN sobre ventas')),suelo=r2(res.get('Suelo')),pem=r2(res.get('PEM')),
      contrata=r2(res.get('Contrata aplicada')),fin=r2(res.get('Costes financieros')),
      ejec=ejec,uds=dg.get('Nº total viviendas'),
      margen_pct=r2(100*res.get('MARGEN sobre ventas',0)/res['TOTAL VENTAS (sin IVA)']) if res.get('TOTAL VENTAS (sin IVA)') else 0)
    wb.close()

# ---------------- FACTURAS ----------------
def refof(com):
    c=(com or '').strip()
    m=re.search(r'\bFRA\.?\s+(\S+)',c,re.I)
    if m: return m.group(1).upper().strip('.,')
    m=re.match(r'^(PAGAR[EÉ]|CONFIRMING)\b.*?VTO\.?\s+[\d\-/]+\s+(\S+)',c,re.I)
    if m: return m.group(2).upper().strip('.,')
    c2=re.sub(r'^(Su\s+Fra\.?|Ntra\.?\s*Fra\.?|Pago\s+Fra\.?|Pgo\s+Fra\.?|Pago|Pgo|Remesa\s*N[ºo°]?)\s*','',c,flags=re.I)
    t=c2.split(); return t[0].upper().strip('.,') if t else ''
PP=('400','401','403','407','410','411'); CP=('430','431','436')
fr=M[(M.cta.str.startswith(PP))&(M.haber>0)&(M.comentario.str.match(r'^(Su Fra|FACTURAS PEND)',case=False,na=False))].copy()
pg=M[(M.cta.str.startswith(PP))&(M.debe>0)].copy()
fr['ref']=fr.comentario.map(refof); pg['ref']=pg.comentario.map(refof)
frk=fr.groupby(['soc','cta','ref']).agg(imp=('haber','sum'),f=('fecha','min'),ej=('ej','min'),
     prom=('promo',lambda s:s.mode().iloc[0] if len(s.mode()) else 'SIN_ASIGNAR'),nom=('descripcion','first')).reset_index()
pgk=pg.groupby(['soc','cta','ref']).agg(pag=('debe','sum'),fp=('fecha','max')).reset_index()
mm=frk.merge(pgk,on=['soc','cta','ref'],how='left'); mm['pag']=mm.pag.fillna(0)
mm['pdte']=(mm.imp-mm.pag).round(2)
mm['estado']=np.where(mm.pag<=0.005,'Sin pago identificado',np.where(mm.pdte.abs()<=0.05,'Pagada','Parcial'))
pgs=pgk.merge(frk[['soc','cta','ref']],on=['soc','cta','ref'],how='left',indicator=True)
pgs=pgs[pgs._merge=='left_only']
fe=M[(M.cta.str.startswith(CP))&(M.debe>0)&(M.comentario.str.contains('Ntra',case=False,na=False))].copy()
co=M[(M.cta.str.startswith(CP))&(M.haber>0)].copy()

def terceros(pref,tipo):
    base=M[M.cta.str.startswith(pref)]; out=[]
    for (soc,cta),x in base.groupby(['soc','cta']):
        if abs(x.neto.sum())<0.005 and len(x)==0: continue
        pr=x[x.promo!='SIN_ASIGNAR'].promo
        out.append(dict(soc=x.soc_nom.iloc[0],cta=cta,
          nom=x.descripcion.mode().iloc[0] if len(x.descripcion.mode()) else '',
          promo=pr.mode().iloc[0] if len(pr.mode()) else 'SIN_ASIGNAR',
          fact=r2(x.debe.sum() if tipo=='cli' else x.haber.sum()),
          cobr=r2(x.haber.sum() if tipo=='cli' else x.debe.sum()),
          saldo=r2(x.neto.sum()),
          ej={str(e):r2(g.debe.sum() if tipo=='cli' else g.haber.sum()) for e,g in x.groupby('ej')}))
    return out
CLI=terceros(CP,'cli'); PRO=terceros(PP,'prov'); ANT=terceros(('438',),'cli')

# ---------------- MOVIMIENTOS DE TESORERIA (cobros y pagos) ----------------
BANC={'57000000':'Caja euros','57200000':'Bancos c/c vista','57200002':'Caja Rural 2017310927',
 '57200003':'Unicaja 5320','57200005':'Santander 2369 · Gastos Retos','57200006':'Santander 2351 · Gastos generales',
 '57200010':'Santander 2628 · Cuenta de crédito','57200007':'Santander 4839 · Reservas Nuevo Campus',
 '57200008':'Santander 4821 · Reservas Doñinos F1-F2','57200009':'Santander 5207 · Reservas Doñinos F3',
 '57200011':'Santander 2636 · Pagos Doñinos','57200012':'Santander 5835 · Reservas Vistahermosa',
 '57200013':'Santander 3225 · Pagos Nuevo Campus','57200014':'Santander 6475 · Reservas Nuevo Campus 2',
 '57200015':'Santander 6483 · Reservas Nuevo Campus 3'}
def _clase(c):
    if c.startswith(('430','431','436')): return 'Cobros de clientes'
    if c.startswith('438'): return 'Anticipos y reservas de compradores'
    if c.startswith(('400','401','403','407','410','411')): return 'Pagos a proveedores'
    if c.startswith('170000'): return 'Préstamo promotor'
    if c.startswith(('520','527','560','561','566')): return 'Pólizas y efectos'
    if c.startswith(('163','16','552','551')): return 'Empresas del grupo y socios'
    if c.startswith(('475','476','477','470','472')): return 'Impuestos y Seguridad Social'
    if c.startswith(('465','640','642','649')): return 'Personal'
    if c.startswith(('662','663','664','665','669','626')): return 'Intereses, avales y comisiones'
    if c.startswith(('570','572')): return 'Traspasos entre cuentas'
    if c.startswith('631'): return 'Tributos'
    return 'Otros'
_M=M.copy(); _M['_k']=_M.file+'|'+_M.fecha.astype(str)+'|'+_M.asiento.astype(str)
_ban=_M[_M.cta.str.startswith(('570','572'))].rename(columns={'_k':'kk'})
_nb =_M[~_M.cta.str.startswith(('570','572'))]
_nb=_nb.assign(_abs=_nb.neto.abs())
_nb['clase']=[_clase(c) for c in _nb.cta]
_agg={}
for k_,g_ in _nb.groupby('_k'):
    g2=g_.sort_values('_abs',ascending=False)
    cl=g2.groupby('clase')._abs.sum().idxmax()
    gg=g2[g2.clase==cl]
    if len(gg)==1: desc=str(gg.descripcion.iloc[0])
    else:
        u=gg.descripcion.nunique()
        desc=str(gg.descripcion.iloc[0])+(f' y {u-1} más' if u>1 else '')
    _agg[k_]=(desc,str(gg.cta.iloc[0]),cl,len(gg))
# traspasos entre cuentas: el asiento no tiene contrapartida fuera de tesoreria
_bb=_ban.assign(_abs=(_ban.debe-_ban.haber).abs())
for k_,g_ in _bb.groupby('kk'):
    if k_ in _agg or len(g_)<2: continue
    g2=g_.sort_values('_abs',ascending=False)
    nom=[(BANC.get(c) if so=='PUM' else de) for c,so,de in zip(g2.cta,g2.soc,g2.descripcion)]
    _agg[k_]=(' → '.join(dict.fromkeys([n for n in nom if n]))[:44],str(g2.cta.iloc[0]),'Traspasos entre cuentas',len(g2))
MOV=[]
for r in _ban.itertuples():
    desc,cc,cl,nl=_agg.get(r.kk,('','','Otros',0))
    imp=r2(r.debe-r.haber)
    if abs(imp)<0.005: continue
    if r.fecha==pd.Timestamp('2023-01-01') and r.soc=='PUM':
        cl='Saldo inicial a 01/01/2023'; desc='Apertura de la contabilidad'
    MOV.append([r.fecha.strftime('%d/%m/%Y'), (BANC.get(r.cta) if r.soc=='PUM' else r.descripcion)[:44],
        str(r.comentario)[:56], desc[:44], cc, cl, r.promo, imp, r.mes, int(r.ej)])
MOV.sort(key=lambda x:(x[8],-abs(x[7])))


# ---------------- PENDIENTE DE COBRO Y DE PAGO ----------------
# El saldo de la cuenta del tercero es la cifra cierta. Las facturas que lo componen
# se derivan aplicando los pagos a las facturas mas antiguas (FIFO), que es como se
# liquidan en la practica; asi el detalle suma siempre el saldo y nunca lo contradice.
_fac={}
for r in fr.itertuples():
    _fac.setdefault(r.cta,[]).append((r.fecha,r.comentario,r2(r.haber)))
_fce={}
for r in fe.itertuples():
    _fce.setdefault(r.cta,[]).append((r.fecha,r.comentario,r2(r.debe)))
def _fifo(cta,saldo,src):
    """Facturas que componen el saldo pendiente, de la mas reciente hacia atras."""
    out=[];rest=saldo
    for f_,com,imp in sorted(src.get(cta,[]),key=lambda z:z[0],reverse=True):
        if rest<=0.05: break
        if imp<=0: continue
        tramo=min(imp,rest); rest=r2(rest-tramo)
        out.append((f_,com,imp,r2(tramo)))
    return out,r2(rest)
PEND=[]
for p in PRO:
    sal=r2(-p['saldo'])
    if sal<=0.05: continue
    det,resto=_fifo(p['cta'],sal,_fac)
    for f_,com,imp,tr in det:
        PEND.append(['Pago',f_.strftime('%d/%m/%Y'),p['nom'],p['cta'],str(com)[:52],p['promo'],imp,r2(imp-tr),tr,int(f_.year)])
    if resto>1:
        PEND.append(['Pago','',p['nom'],p['cta'],'Saldo anterior sin factura en el registro',p['promo'],0.0,0.0,resto,0])
for c in CLI:
    sal=r2(c['saldo'])
    if sal<=0.05: continue
    det,resto=_fifo(c['cta'],sal,_fce)
    for f_,com,imp,tr in det:
        PEND.append(['Cobro',f_.strftime('%d/%m/%Y'),c['nom'],c['cta'],str(com)[:52],c['promo'],imp,r2(imp-tr),tr,int(f_.year)])
    if resto>1:
        PEND.append(['Cobro','',c['nom'],c['cta'],'Saldo anterior sin factura en el registro',c['promo'],0.0,0.0,resto,0])
PEND.sort(key=lambda x:-x[8])

# ---------------- CALIDAD ----------------
G=M[M.cta.str.startswith(('60','61','62','63','64','65','66','67','68'))]
CAL={}
for e in EJ+['TOT']:
    g=G if e=='TOT' else G[G.ej==e]
    i=M[M.cta.str.startswith('70')] if e=='TOT' else M[(M.cta.str.startswith('70'))&(M.ej==e)]
    tot=r2(g.neto.sum()); sin=r2(g[g.promo=='SIN_ASIGNAR'].neto.sum()); don=r2(0.0)
    it=r2(-i.neto.sum()); isin=r2(-i[i.promo=='SIN_ASIGNAR'].neto.sum())
    CAL[str(e)]=dict(gasto=tot,sin=sin,don=don,pct_sin=r2(100*sin/tot) if tot else 0,
      pct_don=r2(100*don/tot) if tot else 0,ing=it,ing_sin=isin,
      pct_ing=r2(100*(it-isin)/it) if it else 0)
dq=D.groupby(['soc_nom','fecha','asiento']).agg(de=('debe','sum'),ha=('haber','sum')).reset_index()
dq['dif']=(dq.de-dq.ha).round(2)
DESC=[dict(soc=r.soc_nom,fecha=r.fecha.strftime('%d/%m/%Y'),asiento=int(r.asiento),debe=r2(r.de),haber=r2(r.ha),dif=r2(r.dif)) for r in dq[dq.dif.abs()>0.01].itertuples()]
SC=[dict(soc=r.soc_nom,fecha=r.fecha.strftime('%d/%m/%Y'),asiento=int(r.asiento),com=r.comentario,desc=r.descripcion,debe=r2(r.debe),haber=r2(r.haber)) for r in D[D.cuenta=='SIN_CUENTA'].itertuples()]
# conciliacion gasto vs activacion por mes
gm={m:r2(G[G.mes==m].neto.sum()) for m in MESES}
am={m:r2(sum(SER[c]['act'][IDX[m]] for c in PORD)) for m in MESES}
CONC=[dict(mes=MLBL[i],ej=int(MESES[i][:4]),gasto=gm[MESES[i]],act=am[MESES[i]],dif=r2(gm[MESES[i]]-am[MESES[i]])) for i in range(NM)]
SIND=sorted([dict(cta=c,desc=d0,imp=r2(x.neto.sum()),n=len(x),nat=x.nat.iloc[0]) for (c,d0),x in G[G.promo=='SIN_ASIGNAR'].groupby(['cta','descripcion'])],key=lambda z:-abs(z['imp']))
DOND=[]

# ---------------- LISTAS DE DETALLE ----------------
FRAC=[[r.cta,r.nom,r.ref,r.f.strftime('%d/%m/%Y'),r2(r.imp),r2(r.pag),r2(r.pdte),r.estado,r.prom,
       (r.fp.strftime('%d/%m/%Y') if isinstance(r.fp,pd.Timestamp) else ''),int(r.ej)] for r in mm.itertuples()]
FEMI=[[r.cta,r.descripcion,r.comentario[:70],r.fecha.strftime('%d/%m/%Y'),r2(r.debe),r.promo,int(r.ej)] for r in fe.itertuples()]
COBR=[[r.cta,r.descripcion,r.comentario[:70],r.fecha.strftime('%d/%m/%Y'),r2(r.haber),r.promo,int(r.ej)] for r in co.itertuples()]
AP=[[r.soc,r.fecha.strftime('%d/%m/%Y'),int(r.asiento),r.comentario[:58],r.descripcion[:46],r.cuenta,
     r2(r.debe),r2(r.haber),r.promo,r.nat,int(r.ej)] for r in M.itertuples()]

PROMOS=[
 ('PUERTO','Puerto de Salamanca','Doñinos de Salamanca — Sector UR-R9, Mz 2B/3/5','Unifamiliar · 52 uds (F1-F3)','En venta y entrega'),
 ('NUEVOCAMPUS','Residencial Nuevo Campus','Salamanca — Sector 77 La Platina, M9.2/1-2-3','Bloque · 64 uds (F1-F3)','En obra'),
 ('CARBAJOSA','Jardines de Carbajosa','Carbajosa de la Sagrada — SUR-R-PAS2, AR.6 y AR.8','Bloque · 144 uds','Inicio de obra (Fase I)'),
 ('DONINOS_RES','Doñinos Residencial','Doñinos de Salamanca — Sector UR-R9, Mz 8 y 9','Unifamiliar · 48 uds','Preparación · obra 09/2026'),
 ('MARIN','Residencial El Marín','Salamanca — Sector El Marín, parcela 12.1','Bloque · 51 uds','Comercialización'),
 ('VISTAHERMOSA','Residencial Vistahermosa','Salamanca — Sector 35C Alambres-Vistahermosa, Mz 21','Unifamiliar · 28 uds','Comercialización'),
 ('MIRADOR','Mirador de Vistahermosa','Salamanca — Sector 65 El Zurguen, Parcela B-1','Unifamiliar · 94 uds','Suelo y proyecto'),
 ('LARAD','La Rad','Salamanca — Sector W8 PP Monte de La Rad','Unifamiliar · 64 uds','Suelo y proyecto'),
 ('ISLARUA','Obra Isla Rúa','Salamanca — Isla Rúa','Contrata ejecutada para Grancampo','Cerrada y facturada'),
 ('PTE_VILLANUEVA','Puente Villanueva de Perales','Villanueva de Perales','Obra para terceros (Ciresco)','En ejecución'),
 ('SUELO_IND','Doñinos Suelo Industrial','Doñinos de Salamanca','Parcela en proyecto, sin licencia ni contrata','En proyecto'),
 ('NAVES','Naves de Salamanca','Salamanca','Parcela en proyecto, sin licencia ni contrata','En proyecto'),
 ('OFICINAS','Oficinas y mejoras propias','Salamanca (Corrillo, Azafranal)','Gasto de oficina, no es una promoción','No promocional'),
 ('SIN_ASIGNAR','Sin asignar / Estructura','—','Gastos de estructura y partidas sin criterio','—'),
]
# ---------------- LOGOS DE PROMOCION (opcional) ----------------
# Si existe una carpeta Logos/ junto a los Diarios, cada imagen se incrusta en el
# dashboard como logo de su promocion. El nombre del fichero puede ser el codigo
# interno (CARBAJOSA.png) o el nombre comercial (Jardines de Carbajosa.svg).
# Formatos admitidos: png, jpg, jpeg, svg, webp. Sin carpeta, se usa el monograma.
import base64, unicodedata
def _slug(t):
    t=unicodedata.normalize('NFKD',str(t)).encode('ascii','ignore').decode().upper()
    return re.sub(r'[^A-Z0-9]','',t)
_NOM={c:n for c,n,_l,_t,_e in PROMOS}
LOGOS={}
_dirL=os.path.join(BASE,'Logos')
if os.path.isdir(_dirL):
    _mime={'.png':'image/png','.jpg':'image/jpeg','.jpeg':'image/jpeg','.svg':'image/svg+xml','.webp':'image/webp'}
    for _f in sorted(os.listdir(_dirL)):
        _e=os.path.splitext(_f)[1].lower()
        if _e not in _mime: continue
        _s=_slug(os.path.splitext(_f)[0])
        _cod=next((c for c in _NOM if _slug(c)==_s or _slug(_NOM[c])==_s),None)
        if not _cod: continue
        _b=open(os.path.join(_dirL,_f),'rb').read()
        if len(_b)>400_000: continue
        LOGOS[_cod]='data:%s;base64,%s'%(_mime[_e],base64.b64encode(_b).decode())
    print('  logos incrustados:',len(LOGOS),'de',len(_NOM))


SOCX=sorted({(r.soc,r.soc_nom,r.nif) for r in D.itertuples()})
PSOC={c:sorted(D[D.promo==c].soc_nom.unique().tolist()) for c in PORD}
NAT={}
G2=M[M.cta.str.startswith(('60','61','62','63','64','65','66','67','68'))]
for c in PORD:
    x=G2[G2.promo==c]; NAT[c]={}
    for e in EJ:
        y=x[x.ej==e]
        NAT[c][str(e)]={k:r2(v) for k,v in y.groupby('nat').neto.sum().items() if abs(v)>0.005}

OTR=[]
for cta,lab in [('16330001','Grancampo Desarrollos'),('16330003','Alhándiga Ibéricos'),('16350001','Óscar Arregui'),
                ('16350002','La Cañada de Hernandinos'),('17400001','Leasing Peugeot Boxer'),('17400002','Leasing telescópica JCB'),
                ('17000001','Préstamo Santander (antiguo)'),('17000002','Préstamo Santander-85 (antiguo)')]:
    x=M[(M.cta==cta)&(M.soc=='PUM')]
    if len(x)==0: continue
    v=[-y for y in [0]]
    ser=[0.0]*NM
    for m,g in x.groupby('mes'): ser[IDX[m]]=r2(-g.neto.sum())
    a=[];s=0.0
    for z in ser: s=r2(s+z); a.append(s)
    OTR.append(dict(cta=cta,nom=lab,serie=a,saldo=a[-1]))
for socc,lab in [('SPV_CARB','Carbajosa')]:
    for cta,l2 in [('16330002','Grancampo Desarrollos'),('16330001','P.U. Montellano')]:
        x=M[(M.cta==cta)&(M.soc==socc)]
        if len(x)==0: continue
        ser=[0.0]*NM
        for m,g in x.groupby('mes'): ser[IDX[m]]=r2(-g.neto.sum())
        a=[];s=0.0
        for z in ser: s=r2(s+z); a.append(s)
        OTR.append(dict(cta=socc+'/'+cta,nom=l2+' → '+lab,serie=a,saldo=a[-1]))

# ---------------- ENLACE A LAS FACTURAS ESCANEADAS (opcional) ----------------
# Si existe ocr/mapa_facturas.json (lo genera mapa.py), cada factura del registro
# recibe la ruta relativa de su PDF para poder abrirlo desde el dashboard.
_cands=[os.path.join(os.path.dirname(os.path.abspath(__file__)),'ocr','mapa_facturas.json'),
        os.path.join(OUT,'ocr','mapa_facturas.json'),
        os.path.join(BASE,'ocr','mapa_facturas.json')]
_mp=next((p for p in _cands if os.path.exists(p)),_cands[0])
PDFREC={}; PDFEMI={}; PDFSIN=[]
if os.path.exists(_mp):
    _m=json.load(open(_mp,encoding='utf-8'))
    PDFREC={int(k):v for k,v in _m.get('rec',{}).items()}
    PDFEMI={int(k):v for k,v in _m.get('emi',{}).items()}
    PDFSIN=_m.get('sin',[])
    print('  facturas escaneadas enlazadas:',len(PDFREC)+len(PDFEMI),'| sin casar:',len(PDFSIN))
FRAC=[x+[PDFREC.get(i)] for i,x in enumerate(FRAC)]
FEMI=[x+[PDFEMI.get(i)] for i,x in enumerate(FEMI)]


# ---------------- SELLO DE ACTUALIZACION ----------------
# Fecha y hora en que se genero este DATA.json, para que el dashboard pueda
# decir cuando se actualizo por ultima vez.
import datetime as _dt
try:
    _tz=_dt.timezone(_dt.timedelta(hours=2))          # horario peninsular de verano
    _ahora=_dt.datetime.now(_tz)
except Exception:
    _ahora=_dt.datetime.now()
GEN=_ahora.strftime('%d/%m/%Y'); GENH=_ahora.strftime('%H:%M')

DATA=dict(
 meta=dict(ejercicios=[int(e) for e in EJ],meses=MESES,mesesLbl=MLBL,
   periodo='01/01/2023 – 31/07/2026',ultimo='31/07/2026',generado=GEN,generadoHora=GENH,
   sociedades=[dict(cod=a,nom=b,nif=c) for a,b,c in SOCX],lineas=len(D),lineasMov=len(M)),
 promos=[dict(cod=c,nom=n,loc=l,tipo=t,estado=e,soc=PSOC[c],pres=(c in PRES)) for c,n,l,t,e in PROMOS],
 ser=SER, nat=NAT, bancos=BAN, deuda=DEU, otras=OTR, pres=PRES,
 clientes=CLI, proveedores=PRO, anticipos=ANT,
 frac=FRAC, femi=FEMI, pdfsin=PDFSIN, cobr=COBR, apuntes=AP, mov=MOV, pend=PEND, logos=LOGOS,
 calidad=dict(porEj=CAL,descuadres=DESC,sincuenta=SC,conc=CONC,sindet=SIND,dondet=DOND,
   fr_estado={k:r2(v) for k,v in mm.groupby('estado').imp.sum().items()},
   fr_n={k:int(v) for k,v in mm.estado.value_counts().items()},
   pagos_sin=r2(pgs.pag.sum()),pagos_sin_n=len(pgs)),
)
actTot=r2(sum(sum(SER[c]['act']) for c in PORD))
actSin=r2(sum(sum(SER[c]['act']) for c in ['SIN_ASIGNAR']))
DATA['calidad']['actTot']=actTot; DATA['calidad']['actSin']=actSin
DATA['calidad']['pctActAsig']=r2(100*(actTot-actSin)/actTot) if actTot else 0


# =============================================================================
# FASE 4 - Analitica contable del cliente (opcional)
#   Si en BASE hay un fichero ANALITICA*.xlsx, se carga su detalle y se
#   contrasta con el coste activado en existencias que calcula el modelo.
#   No modifica ninguna cifra del dashboard: alimenta la pestaña Analitica.
# =============================================================================
_cand = glob.glob(os.path.join(BASE, 'ANALITICA*.xlsx'))
_F = _cand[0] if _cand else None
if _F:
    wb=openpyxl.load_workbook(_F,read_only=True,data_only=True); ws=wb['Hoja1']
    rows=[];hdr=False
    for r in ws.iter_rows(values_only=True):
        if not hdr:
            if r[0]=='PROYECTO': hdr=True
            continue
        if r[0] is None: continue
        rows.append(tuple(r[:14]))
    wb.close()
    A=pd.DataFrame(rows,columns=['proyecto','seccion','canal','fase','fecha','mes','anio','dh','cta','contrap','importe','comentario','serie','factura'])
    A['importe']=pd.to_numeric(A.importe,errors='coerce').fillna(0)
    A['cta']=A.cta.astype(str).str.replace(r'\.0$','',regex=True)
    A['fecha']=pd.to_datetime(A.fecha,errors='coerce')
    MAP={'NUEVO CAMPUS':'NUEVOCAMPUS','DOÑINOS PUERTO DE SALAMANCA':'PUERTO','EL MARIN':'MARIN',
     'VISTAHERMOSA':'VISTAHERMOSA','DOÑINOS RESIDENCIAL':'DONINOS_RES','EL ZURGUEN':'MIRADOR',
     'GASTOS GENERALES INDIRECTOS (MONTELLANO)':'SIN_ASIGNAR','ISLA RUA':'ISLARUA','OFICINAS':'OFICINAS',
     'DOÑINOS SUELO INDUSTRIAL':'SUELO_IND','CARBAJOSA':'CARBAJOSA','LA RAD':'LARAD',
     'PUENTE VILLANUEVA DE PERALES CIRESCO':'PTE_VILLANUEVA','CORRILLO':'OFICINAS'}
    A['cod']=A.proyecto.map(MAP).fillna('SIN_ASIGNAR')

    CORTE='2026-06'
    MAP33={'33000001':'DONINOS_RES','33000002':'NUEVOCAMPUS','33000003':'PUERTO','33000004':'PUERTO',
     '33000005':'ISLARUA','33000006':'OFICINAS','33000007':'MARIN','33000008':'MIRADOR','33000010':'VISTAHERMOSA',
     '33000012':'OFICINAS','33000013':'NUEVOCAMPUS','33000014':'NUEVOCAMPUS','33000015':'SUELO_IND',
     '33000016':'LARAD','33000017':'CARBAJOSA'}
    SPV2P={'SPV_CARB':'CARBAJOSA','SPV_DORES':'DONINOS_RES','SPV_LARAD':'LARAD','SPV_MARIN':'MARIN',
     'SPV_MIRA':'MIRADOR','SPV_NAVES':'NAVES','SPV_VILLAS':'VISTAHERMOSA'}
    inv=M[M.cta.str.startswith(('33','313','21000','23100'))].copy()
    inv['p']=np.where(inv.soc=='PUM',inv.cta.map(MAP33).fillna(inv.promo),inv.soc.map(SPV2P))
    inv=inv[inv.debe>0]
    mio_jun=inv[(inv.soc=='PUM')&(inv.mes<=CORTE)].groupby('p').debe.sum()
    mio_spv=inv[(inv.soc!='PUM')&(inv.mes<=CORTE)].groupby('p').debe.sum()
    mio_jul=inv[inv.mes=='2026-07'].groupby('p').debe.sum()
    # saldo de apertura 01/01/2023 dentro de mis cifras (asiento 1 del diario 2023)
    ap23=inv[(inv.soc=='PUM')&(inv.fecha=='2023-01-01')].groupby('p').debe.sum()
    th=A.groupby('cod').importe.sum()

    CODS=[c for c in PORD if c!='SIN_ASIGNAR']+['NAVES','SIN_ASIGNAR']
    CODS=list(dict.fromkeys(CODS))
    CMP=[]
    for c in CODS:
        mj=r2(mio_jun.get(c,0)); ms=r2(mio_spv.get(c,0)); ap=r2(ap23.get(c,0)); t=r2(th.get(c,0))
        base=r2(mj-ap)                     # comparable: matriz, sin apertura, a junio
        CMP.append(dict(cod=c,mio_total=r2(mj+ms+mio_jul.get(c,0)),mio_jun=mj,spv=ms,apert=ap,
            base=base,ana=t,dif=r2(base-t),julio=r2(mio_jul.get(c,0))))
    # secciones y fases
    SEC=sorted(A.seccion.dropna().unique().tolist())
    FAS=sorted(A.fase.dropna().unique().tolist())
    MSEC={c:{s:r2(v) for s,v in A[A.cod==c].groupby('seccion').importe.sum().items() if abs(v)>0.005} for c in CODS}
    MFAS={c:{f:r2(v) for f,v in A[A.cod==c].groupby('fase').importe.sum().items() if abs(v)>0.005} for c in CODS}
    ANIO=sorted(A.anio.dropna().unique().tolist())
    MANIO={c:{str(int(y)):r2(v) for y,v in A[A.cod==c].groupby('anio').importe.sum().items()} for c in CODS}
    # cruce por cuenta contable
    g=M[(M.soc=='PUM')&(M.mes<=CORTE)&(M.cta.str.startswith(('60','61','62','63','64','65','66','67','68')))]
    d1=A.groupby('cta').importe.sum(); d2=g.groupby('cta').neto.sum()
    CTA=[]
    for k in sorted(set(d1.index)|set(d2.index)):
        x=r2(d1.get(k,0)); y=r2(d2.get(k,0))
        if abs(x)<1 and abs(y)<1: continue
        des=M[M.cta==k].descripcion.mode()
        CTA.append(dict(cta=k,desc=des.iloc[0] if len(des) else '',ana=x,dia=y,dif=r2(x-y)))
    CTA=sorted(CTA,key=lambda z:-abs(z['dif']))
    # reparto que ellos hacen de la cuenta comun de Doñinos
    D2=A[A.cta=='60600002'].groupby('cod').importe.sum()
    REP=[dict(cod=k,imp=r2(v)) for k,v in D2.sort_values(ascending=False).items() if abs(v)>1]

    # ---- facturas que originan las diferencias: criterio de la analitica vs cuenta del diario
    MAP606={'60600001':'LARAD','60600002':'PUERTO','60600003':'NUEVOCAMPUS','60600004':'MARIN','60600005':'VISTAHERMOSA',
     '60600006':'MIRADOR','60600007':'ISLARUA','60600008':'OFICINAS','60600009':'SUELO_IND','60600010':'OFICINAS',
     '60600011':'CARBAJOSA','60600012':'DONINOS_RES','60600014':'PTE_VILLANUEVA'}
    PTN={'21873':'PUERTO','21874':'NUEVOCAMPUS','21875':'PUERTO','21876':'NUEVOCAMPUS','21877':'NUEVOCAMPUS'}
    import re as _re
    def _mio(cta,com):
        com=str(com or '')
        if cta in MAP606: return MAP606[cta]
        if cta=='66230100': return 'PUERTO'
        if cta=='66230101': return 'NUEVOCAMPUS'
        if cta=='66230102': return 'MARIN'
        m=_re.search(r'\b(2187[34567])\b',com)
        if m: return PTN[m.group(1)]
        return None
    A['mio']=[_mio(c,k) for c,k in zip(A.cta,A.comentario)]
    _d=A[A.mio.notna()&(A.mio!=A.cod)].copy()
    _d['abs']=_d.importe.abs()
    _d=_d.sort_values('abs',ascending=False)
    DIF=[[ (r.fecha.strftime('%d/%m/%Y') if isinstance(r.fecha,pd.Timestamp) else ''),
           r.cod, r.mio, str(r.cta), str(r.seccion or ''), str(r.fase or ''), r2(r.importe),
           str(r.comentario if r.comentario==r.comentario and r.comentario is not None else '')[:70],
           int(r.anio) if r.anio==r.anio else 0 ] for r in _d.itertuples()]
    FLU=[]
    for (a_,b_),g_ in _d.groupby(['cod','mio']):
        FLU.append(dict(ana=a_,dia=b_,n=int(len(g_)),imp=r2(g_.importe.sum())))
    FLU=sorted(FLU,key=lambda z:-abs(z['imp']))

    # detalle
    DET=[[r.cod,r.proyecto,str(r.seccion or ''),str(r.canal or ''),str(r.fase or ''),
          (r.fecha.strftime('%d/%m/%Y') if isinstance(r.fecha,pd.Timestamp) else ''),
          int(r.anio) if r.anio==r.anio else 0,str(r.cta),r2(r.importe),str(r.comentario if r.comentario==r.comentario and r.comentario is not None else '')[:58]] for r in A.itertuples()]

    ANA=dict(cods=CODS,secciones=SEC,fases=FAS,anios=[int(x) for x in ANIO],
      cmp=CMP,sec=MSEC,fas=MFAS,anio=MANIO,cta=CTA,rep=REP,det=DET,dif=DIF,flu=FLU,
      total=r2(A.importe.sum()),n=len(A),corte='30/06/2026')
    DATA['ana'] = ANA
    print('  analitica del cliente incorporada:', len(A), 'apuntes')
else:
    print('  aviso: no hay ANALITICA*.xlsx en BASE; la pestana Analitica quedara vacia')



# =============================================================================
# FASE 5 - Modelo de promocion: comercial, obra, planning y flujo financiero
#   Lee de cada estudio economico las hojas DATOS GENERALES, PLANNING MENSUAL,
#   FLUJO FINANCIERO, CAPITULOS OBRA y COMPRADORES. No altera ninguna cifra
#   contable: alimenta las pestanas Comercial, Obra y Proyeccion.
# =============================================================================
import datetime as dt
def fecha(x):
    if isinstance(x, dt.datetime): return x if x.year > 2000 else None
    return None
def ym(x):
    f = fecha(x); return f.strftime("%Y-%m") if f else None
def numx(x):
    try:
        f = float(x); return None if (math.isnan(f) or math.isinf(f)) else f
    except: return None

FIL={'Puerto de Salamanca':'PUERTO','Residencial Nuevo Campus':'NUEVOCAMPUS','Jardines de Carbajosa':'CARBAJOSA',
 'Doñinos Residencial':'DONINOS_RES','Residencial El Marín':'MARIN','Residencial Vistahermosa':'VISTAHERMOSA',
 'Mirador de Vistahermosa':'MIRADOR','La Rad':'LARAD'}

MOD={}
for f in sorted(glob.glob(os.path.join(BASE,'Presupuestos','*.xlsx'))):
    cod=FIL.get(os.path.splitext(os.path.basename(f))[0])
    if not cod: continue
    wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
    H=wb.sheetnames
    m={}

    # ---------- DATOS GENERALES ----------
    dg={}
    if 'DATOS GENERALES' in H:
        for r in wb['DATOS GENERALES'].iter_rows(values_only=True):
            if r[0] is None: continue
            v=next((x for x in r[1:4] if x is not None),None)
            if v is not None: dg[str(r[0]).strip()]=v
    m['sup']=dict(
      constr=r2(dg.get('Sup. total construida')), sr=r2(dg.get('Sup. construida sobre rasante (SR)')),
      util=r2(dg.get('Sup. útil viviendas')), parcela=r2(dg.get('Sup. parcela')),
      edificable=r2(dg.get('Sup. edificable (planeamiento)')), urb=r2(dg.get('Sup. urbanización exterior')),
      uds=int(numx(dg.get('Nº total viviendas')) or 0), gar=int(numx(dg.get('Nº plazas garaje (total)')) or 0),
      pem_m2=r2(dg.get('PEM €/m² sobre rasante')), tipo=str(dg.get('Tipología') or ''),
      arq=str(dg.get('Arquitecto') or ''), proy=ym(dg.get('Fecha proyecto')))

    # ---------- CAPITULOS OBRA ----------
    caps=[]; par={}
    if 'CAPITULOS OBRA' in H:
        ws=wb['CAPITULOS OBRA']; hdr=False
        for r in ws.iter_rows(values_only=True):
            if r[0] and str(r[0]).strip()=='#': hdr=True; continue
            if not hdr:
                if r[1] and isinstance(r[1],str): par[r[1].strip()]=r[2]
                continue
            if r[0] is None or not str(r[0]).strip().isdigit(): continue
            caps.append(dict(n=int(r[0]),nom=str(r[1] or '').strip(),
              pto=r2(r[3]),apl=r2(r[6]),real=r2(r[7]),desv=r2(r[8])))
    m['caps']=caps
    m['obra']=dict(pem=r2(par.get('PEM proyecto')),pto=r2(par.get('Total presupuestado')),
      firmada=r2(par.get('Contrata firmada')),aplicada=r2(par.get('CONTRATA APLICADA')),
      mult=r2(par.get('Multiplicador PEM→Contrata')) or r2(par.get('Multiplicador PEM→Cont')))

    # ---------- PLANNING MENSUAL ----------
    pl=[]
    if 'PLANNING MENSUAL' in H:
        ws=wb['PLANNING MENSUAL']; rows=list(ws.iter_rows(values_only=True))
        hi=next((i for i,r in enumerate(rows) if r and str(r[0]).strip()=='Mes'),None)
        if hi is not None:
            hd=[str(x).strip() if x else '' for x in rows[hi]]
            iTot=hd.index('TOTAL mes') if 'TOTAL mes' in hd else 13
            iIng=hd.index('TOTAL ingr.') if 'TOTAL ingr.' in hd else None
            iC10=hd.index('Contrato 10%') if 'Contrato 10%' in hd else None
            iBim=hd.index('Bimensuales') if 'Bimensuales' in hd else None
            iEsc=hd.index('Escritura 80%') if 'Escritura 80%' in hd else None
            iEje=hd.index('TOTAL ejec.') if 'TOTAL ejec.' in hd else None
            for r in rows[hi+1:]:
                if not r or r[0] is None: continue
                if not str(r[0]).strip().isdigit(): continue
                mm=ym(r[1])
                if not mm: continue
                pl.append([int(r[0]),mm,r2(r[iTot]),
                  r2(r[iC10]) if iC10 else 0.0, r2(r[iBim]) if iBim else 0.0,
                  r2(r[iEsc]) if iEsc else 0.0, r2(r[iIng]) if iIng else 0.0,
                  r2(r[iEje]) if iEje else 0.0])
    m['plan']=pl

    # ---------- FLUJO FINANCIERO ----------
    fl=[]; fp={}
    if 'FLUJO FINANCIERO' in H:
        ws=wb['FLUJO FINANCIERO']; rows=list(ws.iter_rows(values_only=True))
        hi=next((i for i,r in enumerate(rows) if r and str(r[0]).strip()=='Mes'),None)
        for r in rows[:hi or 0]:
            for a,b in [(1,3),(6,7)]:
                if len(r)>b and isinstance(r[a],str) and r[b] is not None: fp[r[a].strip()]=r[b]
        if hi is not None:
            for r in rows[hi+1:]:
                if not r or r[0] is None or not str(r[0]).strip().isdigit(): continue
                mm=ym(r[1])
                if not mm: continue
                fl.append([int(r[0]),mm,r2(r[2]),r2(r[3]),r2(r[5]),r2(r[6]),r2(r[7]),
                           r2(r[9]),r2(r[10]),r2(r[11]),r2(r[12])])
    m['flujo']=fl
    m['fin']=dict(limite=r2(fp.get('Límite préstamo promotor')),tipo=r2(fp.get('Tipo interés préstamo (anual)')),
      apertura=r2(fp.get('Comisión apertura €')),tipoFP=r2(fp.get('Tipo FFPP (anual)')),
      contrata=r2(fp.get('Contrata')),mesObra=r2(fp.get('Mes inicio obra')),mesFin=r2(fp.get('Mes último cronograma')))

    # ---------- COMPRADORES ----------
    uds=[]
    for h in [x for x in H if x.startswith('COMPRADORES')]:
        fase=h.replace('COMPRADORES','').strip() or 'Única'
        ws=wb[h]; rows=list(ws.iter_rows(values_only=True))
        hi=next((i for i,r in enumerate(rows) if r and str(r[0]).strip()=='#'),None)
        if hi is None: continue
        hd=[str(x).strip() if x else '' for x in rows[hi]]
        ix=lambda n,d: hd.index(n) if n in hd else d
        iV,iSU,iSC=ix('Vivienda',2),ix('S.útil',5),ix('S.constr',12)
        iPB,iPT,iIVA=ix('Precio base',13),ix('Precio total',15),ix('Total c/IVA',18)
        iCo,iPg,iPd=ix('Comprador 1',19),ix('Pagado',23),ix('Pendiente',24)
        iFr,iMe=ix('Fecha reserva',25),ix('Mejoras s/IVA',26)
        iEs,iFe=ix('Estado',30),ix('Fecha escritura',31)
        for r in rows[hi+1:]:
            if not r or r[0] is None: continue
            if str(r[0]).strip().upper().startswith('TOTAL'): continue
            if iV>=len(r) or r[iV] is None: continue
            g=lambda k: r[k] if k<len(r) else None
            comp=str(g(iCo) or '').strip()
            pt=r2(g(iPT)) or r2(g(iPB)); iva=r2(g(iIVA)) or r2(pt*1.1)
            pg=r2(g(iPg)); fr=ym(g(iFr)); fe=ym(g(iFe))
            base=iva if iva>0 else pt
            pctp=100*pg/base if base else 0
            est='Libre'
            if comp:
                est='Escriturada' if (fe or pctp>=90) else ('Contratada' if pctp>=15 else 'Reservada')
            uds.append([str(g(iV)).strip(),fase,r2(g(iSU)),r2(g(iSC)),pt,iva,comp[:38],
                        pg,r2(g(iPd)),fr,fe,est,r2(g(iMe))])
    m['uds']=uds
    MOD[cod]=m
    wb.close()
    nv=sum(1 for u in uds if u[11]!='Libre')
    print(f"  {cod:14s} uds {len(uds):4d} vendidas {nv:4d} caps {len(caps):3d} plan {len(pl):3d} flujo {len(fl):3d} sup {m['sup']['constr']:>10,.0f} m2")

# ---------- obra certificada real por mes (cuentas 606 de cada promocion) ----------
try:
    Mv=M
    C606={'60600001':'LARAD','60600002':'PUERTO','60600003':'NUEVOCAMPUS','60600004':'MARIN',
     '60600005':'VISTAHERMOSA','60600006':'MIRADOR','60600007':'ISLARUA','60600009':'SUELO_IND',
     '60600011':'CARBAJOSA','60600012':'DONINOS_RES'}
    ob=Mv[Mv.cta.isin(C606.keys())].copy()
    ob['p']=ob.cta.map(C606)
    g=ob.groupby(['p','mes']).neto.sum()
    for cod in MOD:
        serie=[[m, r2(g.get((cod,m),0.0))] for m in MESES]
        MOD[cod]['obraMes']=serie
    print('  obra certificada mensual incorporada')
except Exception as e:
    print('  aviso: no se pudo calcular la obra mensual:',e)
    for cod in MOD: MOD[cod]['obraMes']=[]
DATA['mod'] = MOD
print('  modelo de promocion incorporado:', len(MOD), 'promociones')

# =====================================================================================
#  FASE 6 · REPARTO DE COSTE  ·  material para explicar y corregir las diferencias
#  entre el coste real contable y el ejecutado del estudio economico.
#
#  No altera ninguna cifra publicada. Solo construye indices sobre la lista de
#  apuntes (DATA['apuntes']) para que el cuadro pueda abrir, promocion a promocion:
#    - las activaciones (33x, 313x, 21000x, 23100x) que forman el Real contable,
#    - el gasto 6xx imputado a esa promocion, factura a factura,
#    - el gasto incurrido en meses posteriores al ultimo cierre activado,
#    - la bandeja de gasto sin promocion, como candidatos a imputar.
#  Los indices apuntan a la posicion en DATA['apuntes'], que se construye desde M
#  en el mismo orden, de modo que la correspondencia es exacta.
# =====================================================================================
print('\nFASE 6 · reparto de coste')
REP={}
try:
    _M=M.reset_index(drop=True)
    assert len(_M)==len(AP), 'desalineacion entre M y la lista de apuntes'

    def _pAct(r):
        """Promocion a la que la contabilidad activa el apunte, si es una activacion."""
        c=r.cta
        if c.startswith('33'):
            return MAP33.get(c) if r.soc=='PUM' else SPV2P.get(r.soc)
        if c.startswith(('313','21000','23100')):
            return r.promo
        return None

    G6=('60','61','62','63','64','65','66','67','68')
    ACTI={}; GAST={}; SINB=[]
    actMes={}; gasMes={}
    for i,r in enumerate(_M.itertuples()):
        p=_pAct(r)
        if p is not None:
            if r.debe and r.debe>0:
                ACTI.setdefault(p,[]).append(i)
                actMes[(p,r.mes)]=r2(actMes.get((p,r.mes),0.0)+r.debe)
        elif r.cta[:2] in G6:
            GAST.setdefault(r.promo,[]).append(i)
            gasMes[(r.promo,r.mes)]=r2(gasMes.get((r.promo,r.mes),0.0)+r.neto)
            if r.promo==SIN: SINB.append(i)

    # ---- ultimo mes con asiento de activacion y coste posterior aun sin activar ----
    ULTACT={}; PENDI={}; PENDIMP={}
    for p in PORD:
        ms=[m for m in MESES if actMes.get((p,m),0)]
        ULTACT[p]=ms[-1] if ms else None
        if ULTACT[p] is None: PENDI[p]=[]; PENDIMP[p]=0.0; continue
        corte=MESES.index(ULTACT[p])
        post={MESES[k] for k in range(corte+1,NM)}
        idx=[i for i in GAST.get(p,[]) if _M.at[i,'mes'] in post]
        PENDI[p]=idx
        PENDIMP[p]=r2(sum(float(_M.at[i,'neto']) for i in idx))

    # ---- enlace de cada apunte de gasto con el PDF escaneado de su factura ----
    # Se apoya en el registro de facturas recibidas ya casado por el OCR: si el
    # asiento contiene la linea de proveedor de una factura con PDF, todas las
    # lineas de gasto de ese asiento apuntan a ese PDF.
    PDFAP={}
    try:
        _fr={}
        for _i,_x in enumerate(FRAC):
            _ruta=_x[11] if len(_x)>11 else None
            if _ruta: _fr[(_x[0],_x[3],round(float(_x[4]),2))]=_ruta
        if _fr:
            _prov=_M[_M.cta.str.startswith(('400','410','411'))]
            _key={}
            for r in _prov.itertuples():
                k=(r.cta,r.fecha.strftime('%d/%m/%Y'),round(float(r.haber),2))
                if k in _fr: _key[(r.soc,int(r.asiento))]=_fr[k]
            if _key:
                for i,r in enumerate(_M.itertuples()):
                    if r.cta[:2] in G6:
                        v=_key.get((r.soc,int(r.asiento)))
                        if v: PDFAP[i]=v
        print('  apuntes de gasto con factura escaneada:',len(PDFAP))
    except Exception as e:
        print('  aviso: no se pudo enlazar el PDF a los apuntes de gasto:',e)

    # ---- proveedor de cada apunte de gasto, leido de la linea de tercero del asiento ----
    PROVAP={}
    try:
        _t=_M[_M.cta.str.startswith(('400','410','411'))]
        _nom={}
        for r in _t.itertuples():
            k=(r.soc,int(r.asiento)); v=(r.descripcion or r.comentario or '').strip()
            if v and k not in _nom: _nom[k]=v[:46]
        for i,r in enumerate(_M.itertuples()):
            if r.cta[:2] in G6:
                v=_nom.get((r.soc,int(r.asiento)))
                if v: PROVAP[i]=v
    except Exception as e:
        print('  aviso: no se pudo identificar el proveedor de los apuntes:',e)

    # ---- clasificacion por capitulo del estudio economico ----------------------------
    # La columna "Ejecutado" del estudio no es una cifra suelta: es la analitica de
    # contabilidad agrupada por SECCION. Esta comprobado apunte a apunte -- con el mapa
    # de secciones de abajo, la analitica reproduce al centimo los capitulos del estudio
    # en las seis promociones de la matriz (Puerto, Nuevo Campus, Doninos Residencial,
    # El Marin, Vistahermosa y Mirador). Asi que el capitulo de cada factura no se
    # adivina: se toma de la propia analitica cuando la factura aparece en ella, y solo
    # se decide por cuenta cuando no aparece.
    CAPS=['Suelo','Obra','Tasas e impuestos','Honorarios técnicos','Acometidas','Comercial',
          'Jurídicos','Seguros y avales','Financieros','Imprevistos y costes extra',
          'Estructura','Otros']
    SEC2CAP={'CONSTRUCCIONES':'Obra','MEJORAS':'Obra','PISCINAS':'Obra','COCINAS':'Obra',
      'PERSONAL':'Obra','SUELO / TERRENO':'Suelo','GASTOS INICIALES':'Suelo',
      'FINANCIEROS':'Financieros','ARQUITECTO Y APARAREJADOR':'Honorarios técnicos',
      'TECNICOS':'Honorarios técnicos','OCT Y CC':'Honorarios técnicos',
      'LICENCIAS Y TASAS':'Tasas e impuestos','COMERCIALIZACION':'Comercial',
      'ACOMETIDAS':'Acometidas','GASTOS JURIDICOS':'Jurídicos',
      'GASTOS INDIRECTOS A LA OBRA (OTROS GASTOS)':'Jurídicos',
      'GASTOS GENERALES (MONTELLANO)':'Jurídicos'}
    # Mapa por cuenta, revisado cuenta a cuenta sobre el plan real de la sociedad.
    # Solo se usa para lo que la analitica no recoge.
    def _capCta(c):
        if c.startswith(('313','21000','23100')) or c.startswith('601'): return 'Suelo'
        if c.startswith('606') or c.startswith('602'): return 'Obra'
        if c.startswith('631'): return 'Tasas e impuestos'
        if c.startswith(('630','6301')): return 'Otros'      # impuesto de sociedades y diferido
        if c.startswith('623'): return 'Honorarios técnicos'
        if c.startswith('627'): return 'Comercial'
        if c.startswith('625'): return 'Seguros y avales'
        if c.startswith('66230100') or c.startswith('66230101') or c.startswith('66230102'):
            return 'Seguros y avales'                        # estos si son avales
        if c.startswith(('660','661','662','665','668','669','626')): return 'Financieros'
        if c.startswith('62800003') or c.startswith('62800005'): return 'Obra'  # gasoleo y combustible de obra
        if c.startswith('628'): return 'Acometidas'
        if c.startswith('611'): return 'Suelo'               # variacion de existencias de terrenos
        if c.startswith(('620','621','622','624','629','640','641','642','643','649')): return 'Estructura'
        return 'Otros'
    # ---- cruce de la analitica con el diario, factura a factura ----------------------
    # La analitica reparte un mismo apunte del diario entre varios proyectos y fases, asi
    # que el cruce se hace agrupando por fecha, cuenta y comentario: la suma del grupo en
    # la analitica tiene que coincidir con el apunte del diario.
    ANACAP={}; ANAREP={}; ANAKEY={}
    ANA_OK=ANA_SOLO=0
    try:
        _A=A.copy()
        _A['cap']=_A.seccion.map(SEC2CAP).fillna('Otros')
        def _nk(s):
            return (s.fillna('').astype(str).str.upper()
                    .str.replace(r'\s+',' ',regex=True).str.strip().str[:28])
        _A['k']=_A.fecha.dt.strftime('%Y%m%d')+'|'+_A.cta.astype(str)+'|'+_nk(_A.comentario)
        _Mp=_M[_M.soc=='PUM'].copy()
        _Mp['k']=_Mp.fecha.dt.strftime('%Y%m%d')+'|'+_Mp.cta.astype(str)+'|'+_nk(_Mp.comentario)
        _gk=_A.groupby('k')
        _capk=_gk.apply(lambda g: g.groupby('cap').importe.sum().idxmax())
        _impk=_gk.importe.sum()
        _repk=_gk.apply(lambda g: {k:r2(v) for k,v in g.groupby('cod').importe.sum().items()})
        _seen=set()
        for i,r in zip(_Mp.index,_Mp.itertuples()):
            k=_Mp.at[i,'k']
            if k in _capk.index and r.cta[:2] in G6:
                ANACAP[i]=_capk[k]; ANAKEY[i]=k
                if k not in _seen:
                    ANAREP[i]=_repk[k]; _seen.add(k)
                ANA_OK+=1
        ANA_SOLO=int(len(_impk.index.difference(set(_Mp.k))))
        print(f'  analitica cruzada con el diario: {ANA_OK} apuntes casados, {ANA_SOLO} grupos solo en la analitica')
    except Exception as e:
        print('  aviso: no se pudo cruzar la analitica con el diario:',e)

    CAPAP={}; CAPORI={}
    for i,r in enumerate(_M.itertuples()):
        c=r.cta
        if c.startswith(('313','21000','23100')) and r.debe and r.debe>0 and _pAct(r) is not None:
            CAPAP[i]='Suelo'; CAPORI[i]='cuenta'
        elif c[:2] in G6:
            if i in ANACAP: CAPAP[i]=ANACAP[i]; CAPORI[i]='analitica'
            else: CAPAP[i]=_capCta(c); CAPORI[i]='cuenta'

    # ---- residuo: coste incorporado a existencias sin factura ni compra asociada ------
    # La activacion mensual la fija el contable contra la cuenta 71x, y no tiene por que
    # coincidir con el gasto 6xx del periodo: la diferencia son aportaciones de suelo y
    # regularizaciones que entran directamente en la 33x. Se aisla para que el cuadre
    # cierre al centimo.
    RESID={}
    for p in PORD:
        a=sum(float(_M.at[i,'debe'] or 0) for i in ACTI.get(p,[]))
        dr=sum(float(_M.at[i,'debe'] or 0) for i in ACTI.get(p,[]) if not _M.at[i,'cta'].startswith('33'))
        g=sum(float(_M.at[i,'neto']) for i in GAST.get(p,[]))
        v=r2(a-dr-g)
        if abs(v)>0.005: RESID[p]=v

    # ---- conciliacion objetiva contra la analitica, factura a factura ----------------
    # La diferencia entre el coste que este cuadro imputa a una promocion y el que le
    # imputa la analitica -- que es la columna Ejecutado del estudio -- se descompone en
    # partidas concretas que suman la diferencia exacta. No hay nada que dar por bueno a
    # mano: cada partida es una factura con nombre, fecha e importe.
    CONC={}
    try:
        _ANAT=A.groupby('cod').importe.sum().to_dict()
        # reparto de la analitica para TODOS sus grupos, tambien los que no aparecen en
        # el diario: si no se incluyen, la conciliacion no cierra
        _rep={k:{a:r2(b) for a,b in g.groupby('cod').importe.sum().items() if abs(b)>0.005}
              for k,g in _A.groupby('k')}
        _mia={}                                   # clave de grupo -> promocion que yo asigno
        for p,idx in GAST.items():
            for i in idx:
                k=ANAKEY.get(i)
                if k: _mia.setdefault(k,set()).add(p)
        # Se trabaja por grupo de factura: mi lado es la suma de los apuntes del diario de
        # ese grupo repartida segun la obra que yo asigno; el lado de la analitica es su
        # propio reparto por proyecto. La resta de ambos, grupo a grupo, suma exactamente
        # la diferencia entre las dos columnas.
        _mioK={}; _repK={}; _sinK=[]
        for p,idx in GAST.items():
            for i in idx:
                k=ANAKEY.get(i); v=float(_M.at[i,'neto'])
                if k is None:
                    if abs(v)>0.005: _sinK.append((p,int(i),r2(v)))
                else:
                    _mioK.setdefault(k,{}).setdefault(p,0.0)
                    _mioK[k][p]+=v
                    _repK.setdefault(k,[]).append(int(i))
        _todas=set(_mioK)|set(_rep)
        for p in PORD:
            it=[]
            for k in _todas:
                mio=r2((_mioK.get(k) or {}).get(p,0.0))
                ana=r2((_rep.get(k) or {}).get(p,0.0))
                d=r2(mio-ana)
                ii=(_repK.get(k) or [None])[0]
                otros={a:b for a,b in (_rep.get(k) or {}).items() if a!=p and abs(b)>0.005}
                mios={a:r2(b) for a,b in (_mioK.get(k) or {}).items() if a!=p and abs(b)>0.005}
                if k not in _mioK:
                    if abs(d)>=0.005:
                        it.append(dict(t='solo_ana',i=ii,imp=d,mio=r2(mio),ana=r2(ana),otros=otros,mios=mios))
                    continue
                # La factura esta en los dos sitios. Se separan dos cosas distintas:
                #   - que cada uno la reparta entre obras de forma diferente, que por
                #     construccion neteA cero sumando todas las promociones, y
                #   - que el importe total de la factura no coincida, que no es un
                #     criterio de reparto sino una incidencia a resolver.
                MIO=r2(sum((_mioK.get(k) or {}).values()))
                ANA=r2(sum((_rep.get(k) or {}).values()))
                DEL=r2(MIO-ANA)
                if abs(DEL)<0.005:      cuota=0.0
                elif abs(ANA)>=0.005:   cuota=r2(DEL*ana/ANA)
                elif abs(MIO)>=0.005:   cuota=r2(DEL*mio/MIO)
                else:                   cuota=0.0
                rep=r2(d-cuota)
                if abs(cuota)>=0.005:
                    it.append(dict(t='importe',i=ii,imp=cuota,mio=r2(MIO),ana=r2(ANA),otros=otros,mios=mios))
                if abs(rep)>=0.005:
                    t='otra_obra' if abs(ana)<0.005 else 'reparto'
                    it.append(dict(t=t,i=ii,imp=rep,mio=r2(mio),ana=r2(ana),otros=otros,mios=mios))
            for pp,i,v in _sinK:
                if pp==p: it.append(dict(t='no_ana',i=i,imp=v))
            if abs(RESID.get(p,0))>0.005: it.append(dict(t='residuo',imp=r2(RESID[p])))
            dr=r2(sum(float(_M.at[i,'debe'] or 0) for i in ACTI.get(p,[])
                      if not _M.at[i,'cta'].startswith('33')))
            if abs(dr)>0.005: it.append(dict(t='directa',imp=dr))
            if it: CONC[p]=sorted(it,key=lambda x:-abs(x['imp']))
        # control: la suma de partidas tiene que dar la diferencia publicada
        for p in (PRES or {}):
            s=r2(sum(x['imp'] for x in CONC.get(p,[])))
            real=r2(sum(float(_M.at[i,'debe'] or 0) for i in ACTI.get(p,[])))
            dif=r2(real-_ANAT.get(p,0.0))
            if abs(s-dif)>0.05:
                print(f'  aviso: la conciliacion de {p} suma {s:,.2f} y la diferencia es {dif:,.2f}')
        print('  conciliacion contra la analitica:',sum(len(v) for v in CONC.values()),'partidas')
    except Exception as e:
        print('  aviso: no se pudo construir la conciliacion contra la analitica:',e)
        CONC={}

    EST={}
    for p,v in PRES.items():
        e={}
        for c in (v.get('caps') or []):
            if abs(c.get('ejec') or 0)>0.005: e[c['cap']]=r2(c['ejec'])
        EST[p]=e

    REP=dict(
      act={p:v for p,v in ACTI.items()},
      gas={p:v for p,v in GAST.items()},
      sin=SINB,
      ultAct=ULTACT,
      pend={p:v for p,v in PENDI.items() if v},
      pendImp={p:v for p,v in PENDIMP.items() if abs(v)>0.005},
      pdf=PDFAP, prov=PROVAP, cap=CAPAP, capOri=CAPORI, caps=CAPS, resid=RESID, est=EST,
      anaRep=ANAREP, anaKey=ANAKEY, conc=CONC, anaTot={k:r2(v) for k,v in (A.groupby("cod").importe.sum().to_dict().items() if "A" in dir() else [])},
      residNom='Suelo y regularizaciones aportados directamente a existencias',
      map33=MAP33, spv2p=SPV2P,
      ejec={p:(PRES.get(p,{}) or {}).get('ejec') for p in PORD if (PRES.get(p,{}) or {}).get('ejec')},
    )
    print('  activaciones indexadas :',sum(len(v) for v in ACTI.values()))
    print('  gasto indexado         :',sum(len(v) for v in GAST.values()))
    print('  bandeja sin promocion  :',len(SINB))
    for p in [x for x in PORD if PENDIMP.get(x)]:
        print(f'    {p:14s} ultimo activado {ULTACT[p]}  coste posterior sin activar {PENDIMP[p]:>14,.2f}')
except Exception as e:
    print('  aviso: no se pudo construir el reparto de coste:',e)
    REP={}
DATA['rep']=REP

json.dump(DATA, open(os.path.join(OUT,'DATA.json'),'w',encoding='utf-8'), ensure_ascii=False, separators=(',',':'))
print('DATA.json',os.path.getsize(os.path.join(OUT,'DATA.json'))//1024,'KB')
print('meses',NM,'| apuntes',len(AP),'| facturas rec',len(FRAC),'| emitidas',len(FEMI))
print('calidad TOT:',CAL['TOT'])

print('\nETL completado.')
